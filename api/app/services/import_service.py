# Servicio de importación masiva desde CSV o Excel
import io
from typing import List, Tuple

from sqlalchemy.orm import Session

from app.models.product import Product
from app.services.product_service import product_service


def parse_csv_or_excel(file_bytes: bytes, filename: str) -> List[dict]:
    """Parsea CSV o Excel y devuelve lista de dicts con los datos de cada fila."""
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows
    else:
        import csv
        content = file_bytes.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        return [dict(row) for row in reader]


def bulk_import(db: Session, file_bytes: bytes, filename: str) -> Tuple[int, List[str]]:
    """
    Importa productos en masa desde CSV o Excel.
    Devuelve (count_created, list_of_errors).
    """
    rows = parse_csv_or_excel(file_bytes, filename)
    created = 0
    errors = []
    created_products: list[Product] = []

    for i, row in enumerate(rows, start=2):  # fila 2 = primera de datos
        name = row.get("name") or row.get("nombre")
        if not name:
            errors.append(f"Fila {i}: campo 'name' obligatorio")
            continue

        # Mapeo flexible de columnas en español e inglés
        product = Product(
            name=str(name),
            set_number=row.get("set_number") or row.get("numero_set"),
            theme=row.get("theme") or row.get("tema"),
            condition=row.get("condition") or row.get("condicion") or None,
            purchase_price=_to_decimal(row.get("purchase_price") or row.get("precio_compra")),
            purchase_source=row.get("purchase_source") or row.get("fuente_compra"),
            quantity=int(row.get("quantity") or row.get("cantidad") or 1),
            notes=row.get("notes") or row.get("notas"),
            availability="available",
        )
        db.add(product)
        created_products.append(product)
        created += 1

    db.commit()

    for product in created_products:
        try:
            product_service.enrich_market_history_if_possible(db, product)
        except Exception:
            errors.append(
                f"Producto {product.set_number or product.name}: no se pudo cargar histórico de precios de los últimos 6 meses"
            )

    return created, errors


def _to_decimal(value):
    if value is None or value == "":
        return None
    try:
        from decimal import Decimal
        return Decimal(str(value).replace(",", "."))
    except Exception:
        return None
